import pandas as pd
import openpyxl
import matplotlib as plt
import numpy as np
import os

name = '2 mM'
path = r'G:\TFG\20220411\over 70.xlsx' #output excel
filename = os.path.join(r'G:\TFG\20220411',name + '.xlsx') ##run excel
time = [1,3,5,10,15]
data_mean = 42
data_std = 4.5
row = 4
mode = 1 ##if mode=0,add an excel

result = list()
for t in time:
    sheetname = str(t)
    df = pd.read_excel(filename, engine='openpyxl', sheet_name=sheetname, usecols=[1])
    data_original = np.array(df)
    data = []
    for i in data_original:
        if (i > 0 and i < 120):
            data.append(float(i))
    data.sort()
    n = 0
    for j in data:
        if (j > 70):
            n = n + 1
    percentage = int(n / (len(df) + 1) * 100)
    # percentage = str(percentage) + '%'
    result.append(percentage)
print(result)


##create new excel
if (mode==0):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for i in range(len(time)):
        worksheet.cell(1, i+2,time[i])
    workbook.save(filename=path)
##open excel
if(mode==1):
    number = str("A") + str(row)
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    worksheet[number] = name
    for i in range(len(result)):
        worksheet.cell(row, i+2,result[i])
    workbook.save(filename=path)
