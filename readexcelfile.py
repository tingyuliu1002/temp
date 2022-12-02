import openpyxl as xl
import pandas as pd
from basic.select import select_file
df = pd.read_excel (r'Path where the Excel file is stored\File name.xlsx', sheet_name='avg_attrs')
print (df)

path1 = 'C:\\Users\\Xukrao\\Desktop\\workbook1.xlsx'
path2 = 'C:\\Users\\Xukrao\\Desktop\\workbook2.xlsx'

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)



wb2.save(path2)